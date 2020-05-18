import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

out = cv2.VideoWriter('Preeti_Swati_CaseStudyAssignment2_video.mp4', cv2.VideoWriter_fourcc(*'mp4v'), 30, (360, 288))


class Result:

    def __init__(self):
        self.type_list = list()
        self.colour_list = list()
        self.picture = None


    def update_result(self, frame, frame_n, type, colour, max_car, bound):

        if max_car > 1:
            self.type_list.append(type)
            self.colour_list.append(colour)
            self.picture = self.write_video(frame, max_car, type, colour, bound, self.picture)
        elif max_car == 1:
            img = self.write_video(frame, max_car, type, colour, bound, self.picture)
            self.show_video(frame_n, img)
            self.picture = None

            self.type_list.append(type)
            self.colour_list.append(colour)
            self.write_excel(frame_n, self.type_list, self.colour_list)
            self.type_list.clear()
            self.colour_list.clear()
        elif max_car == 0:
            img = np.array(frame)
            self.show_video(frame_n, img)
            self.write_excel(frame_n, self.type_list, self.colour_list)

    def write_video(self, frame, max_car, type, colour, bound, image1=None):
        # figure;
        if image1 is None:
            image1 = np.array(frame)

        font = cv2.FONT_HERSHEY_PLAIN
        org1 = (150, 60)
        org2 = (bound[0] + 50, bound[1] - 15)
        org3 = (bound[0], bound[1] - 15)
        fontScale = 1
        color = (0, 0, 0)
        thickness = 2
        countstring = "Car Count: " + str(max_car)
        image1 = cv2.putText(image1, countstring, org1, font,
                             fontScale, color, thickness, cv2.LINE_AA)
        image1 = cv2.putText(image1, type, org2, font,
                             fontScale, color, thickness, cv2.LINE_AA)
        image1 = cv2.putText(image1, colour, org3, font,
                             fontScale, color, thickness, cv2.LINE_AA)
        return image1

    def show_video(self, frame_n, img, done=False):
        if not done:
            filename = "C:\\Users\\User\\PycharmProjects\\RTAIAssignment2\\keras-yolo3\\ClassifiedData\\image_" \
                       + str(frame_n) + ".jpg"
            cv2.imshow('Object Detection and Classification', img)
            cv2.imwrite(filename, img)
            out.write(img)
            cv2.waitKey(10)
        else:
            out.release()

    def write_excel(self, frame_num, car_info, colour_info):
        filename = 'CarClassifier.xlsx'
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.merge_cells('A1:A2')
            ws.cell(row=1, column=1).value = 'Frame No.'

            ws.merge_cells('B1:F1')
            ws.cell(row=1, column=2).value = 'Sedan'

            ws.merge_cells('G1:K1')
            ws.cell(row=1, column=7).value = 'Hatchback'

            ws.merge_cells('L1:L2')
            ws.cell(row=1, column=12).value = 'Total Cars'

            ws.cell(row=2, column=2).value = 'Black'
            ws.cell(row=2, column=3).value = 'Silver'
            ws.cell(row=2, column=4).value = 'Red'
            ws.cell(row=2, column=5).value = 'White'
            ws.cell(row=2, column=6).value = 'Blue'

            ws.cell(row=2, column=7).value = 'Black'
            ws.cell(row=2, column=8).value = 'Silver'
            ws.cell(row=2, column=9).value = 'Red'
            ws.cell(row=2, column=10).value = 'White'
            ws.cell(row=2, column=11).value = 'Blue'

        sred, sblue, sblack, swhite, ssilver = 0, 0, 0, 0, 0
        hred, hblue, hblack, hwhite, hsilver = 0, 0, 0, 0, 0

        if len(car_info) > 0:
            for i in range(len(car_info)):
                if colour_info[i] == "Black" and car_info[i] == "Sedan":
                    sblack += 1
                elif colour_info[i] == "Silver" and car_info[i] == "Sedan":
                    ssilver += 1
                elif colour_info[i] == "Red" and car_info[i] == "Sedan":
                    sred += 1
                elif colour_info[i] == "White" and car_info[i] == "Sedan":
                    swhite += 1
                elif colour_info[i] == "Blue" and car_info[i] == "Sedan":
                    sblue += 1
                elif colour_info[i] == "Black" and car_info[i] == "Hatchback":
                    hblack += 1
                elif colour_info[i] == "Silver" and car_info[i] == "Hatchback":
                    hblack += 1
                elif colour_info[i] == "Red" and car_info[i] == "Hatchback":
                    hred += 1
                elif colour_info[i] == "White" and car_info[i] == "Hatchback":
                    hwhite += 1
                elif colour_info[i] == "Blue" and car_info[i] == "Hatchback":
                    hblue += 1

            ws.append([frame_num, sblack, ssilver, sred, swhite, sblue,
                       hblack, hsilver, hred, hwhite, hblue, len(car_info)])
        else:
            ws.append([frame_num, sblack, ssilver, sred, swhite, sblue,
                       hblack, hsilver, hred, hwhite, hblue, len(car_info)])
        wb.save(filename)
